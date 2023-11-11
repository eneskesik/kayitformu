from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/kaydet', methods=['POST'])
def kaydet():
    ad = request.form['ad']
    soyad = request.form['soyad']
    yas = request.form['yas']

    # Ad ve soyadı capitalize et
    ad = ad.capitalize()
    soyad = soyad.capitalize()

    # Excel dosyasına ekle
    dosya = openpyxl.load_workbook("deneme.xlsx")
    sayfa = dosya["Tablo"]
    yeni_sira = sayfa.max_row + 1
    sayfa.cell(row=yeni_sira, column=1, value=ad)
    sayfa.cell(row=yeni_sira, column=2, value=soyad)
    sayfa.cell(row=yeni_sira, column=3, value=yas)
    dosya.save("deneme.xlsx")

    return render_template('index.html', message='Kişi başarıyla eklendi.')


if __name__ == '__main__':
    app.run(debug=True)
